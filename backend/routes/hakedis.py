"""
Hakedis Raporu (Progress Payment Report) Blueprint
Generates aggregated payment reports grouped by project, with current-period
and cumulative columns, then exports to Excel.

Also provides endpoints to upload an Access database (.accdb/.mdb) and extract
its table structure + VBA macros so the app can mirror the original logic.
"""

from flask import Blueprint, jsonify, request, send_file
import json
import io
import os
import re
import tempfile
from datetime import datetime, date, timedelta
from collections import defaultdict
from werkzeug.utils import secure_filename

from models import db
from models.database_record import DatabaseRecord
from middleware.auth_middleware import role_required

hakedis_bp = Blueprint('hakedis', __name__, url_prefix='/api/hakedis')

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), '..', 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)
ALLOWED_ACCESS_EXTENSIONS = {'.accdb', '.mdb'}

# In-memory store for the last parsed Access structure (server-scoped, single user OK)
_access_store = {}   # { 'path': str, 'tables': [...], 'queries': [...], 'macros': [...] }

# In-memory store for uploaded Excel files (one per type)
_excel_store = {}    # { type_key: { 'path': str, 'filename': str, 'uploaded_at': str } }

EXCEL_TYPES = {
    'database':        {'label': 'DATABASE',           'description': 'Ana veri — personel, disiplin, MH, şirket, proje bilgileri'},
    'hourly_rates':    {'label': 'Hourly Rates',       'description': 'Saatlik ücretler — personel ve sözleşme bazlı birim fiyatlar'},
    'hakedis':         {'label': 'Hakedis',            'description': 'Hakediş takibi — sözleşme no, firma, hakediş no, dönem, tutar'},
    'doviz_kurlari':   {'label': 'Döviz Kurları',      'description': 'Döviz kuru tablosu — haftalık USD/TRY, EUR/USD kurları'},
    'sirket_isimleri': {'label': 'Şirket İsimleri',    'description': 'Şirket isim eşleştirme — kısa ad / uzun ad tablosu'},
    'info':            {'label': 'Info',               'description': 'Kapsam, proje ve birim fiyat bilgileri — hesaplama referans tablosu'},
}

def _restore_excel_store():
    """Repopulate _excel_store from files already saved in UPLOAD_DIR (survives backend restarts)."""
    import pandas as pd
    for excel_type in EXCEL_TYPES:
        for ext in ('.xlsx', '.xls', '.xlsb'):
            candidate = os.path.join(UPLOAD_DIR, f'hakedis_{excel_type}{ext}')
            if os.path.exists(candidate):
                row_count = 0
                columns = []
                try:
                    xl = pd.ExcelFile(candidate)
                    df = pd.read_excel(candidate, sheet_name=xl.sheet_names[0])
                    row_count = len(df)
                    columns = list(df.columns)[:20]
                except Exception as e:
                    print(f'[EXCEL STORE RESTORE] Could not read {candidate}: {e}')
                mtime = datetime.fromtimestamp(os.path.getmtime(candidate)).strftime('%Y-%m-%d %H:%M')
                _excel_store[excel_type] = {
                    'path':        candidate,
                    'filename':    os.path.basename(candidate),
                    'uploaded_at': mtime,
                    'row_count':   row_count,
                    'columns':     columns,
                }
                print(f'[EXCEL STORE RESTORE] Restored "{excel_type}" from {candidate}')
                break  # found one extension, skip others

_restore_excel_store()

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

MONTH_MAP = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
}


def _load_records():
    """
    Load personnel/cost records.
    Priority:
      1. Hakedis DATABASE Excel (uploaded via the Hakedis page) — always most up-to-date.
      2. PostgreSQL DatabaseRecord table — fallback if no Excel is uploaded.
    """
    db_store = _excel_store.get('database')
    if db_store and os.path.exists(db_store.get('path', '')):
        try:
            import pandas as pd
            path = db_store['path']
            xl = pd.ExcelFile(path)
            # Prefer a sheet called DATABASE / Database; else use first sheet
            sheet_name = next(
                (s for s in xl.sheet_names if s.strip().upper() == 'DATABASE'),
                xl.sheet_names[0]
            )
            df = pd.read_excel(path, sheet_name=sheet_name, dtype=str)
            df = df.fillna('')
            # Drop rows where every cell is empty
            df = df[df.apply(lambda r: any(v.strip() for v in r.values if isinstance(v, str)), axis=1)]
            records = df.to_dict('records')
            # Normalise: strip whitespace from all string values, replace 'nan'/'NaT'
            _bad = {'nan', 'NaT', 'None', 'none'}
            cleaned = []
            for rec in records:
                cleaned.append({
                    k: ('' if str(v).strip() in _bad else str(v).strip())
                    for k, v in rec.items()
                })
            if cleaned:
                print(f'[_load_records] Using DATABASE Excel: {len(cleaned)} rows from {db_store["filename"]}')
                return cleaned
        except Exception as e:
            print(f'[_load_records] Excel read failed ({e}), falling back to PostgreSQL')

    # Fallback: PostgreSQL
    rows = DatabaseRecord.query.all()
    out = []
    for r in rows:
        try:
            out.append(json.loads(r.data))
        except Exception:
            continue
    print(f'[_load_records] Using PostgreSQL: {len(out)} rows')
    return out


def _safe_str(v):
    if v is None:
        return ''
    s = str(v).strip()
    return '' if s in ('nan', 'NaT', 'None') else s


def _safe_float(v):
    try:
        return float(_safe_str(v))
    except (ValueError, TypeError):
        return 0.0


def _parse_record_date(record):
    """Parse the (Week / Month) field → datetime.date or None.
    Tries all known column name variants created by different Excel exports.
    """
    raw = (_safe_str(record.get('(Week / \nMonth)', '')) or   # space before \n  (new Excel)
           _safe_str(record.get('(Week /\nMonth)', '')) or    # no space before \n
           _safe_str(record.get('(Week / Month)', '')) or
           _safe_str(record.get('(Week /\n Month)', '')) or
           _safe_str(record.get('Week / Month', '')) or
           _safe_str(record.get('Week/Month', '')) or
           _safe_str(record.get('Week/\nMonth', '')))
    if not raw:
        return None
    # Try DD/Mon/YYYY  e.g.  01/Dec/2024
    try:
        return datetime.strptime(raw, '%d/%b/%Y').date()
    except ValueError:
        pass
    # Try YYYY-MM-DD
    try:
        return datetime.strptime(raw[:10], '%Y-%m-%d').date()
    except ValueError:
        pass
    # Try DD.MM.YYYY
    try:
        return datetime.strptime(raw[:10], '%d.%m.%Y').date()
    except ValueError:
        pass
    return None


def _get_mh(record):
    """Read TOTAL MH from a record, trying all known column name variants.
    VBA uses TOTAL_MH (underscore) first, then TOTAL MH (space).
    """
    return (_safe_float(record.get('TOTAL_MH')) or        # VBA: first choice (underscore)
            _safe_float(record.get('TOTAL\n MH')) or
            _safe_float(record.get('TOTAL MH')) or
            _safe_float(record.get('Total MH')) or
            _safe_float(record.get('TOTAL\nMH')))


def _get_kapsam(record):
    """Read the İşveren- Hakediş Kapsam scope key, trying column name variants."""
    return (_safe_str(record.get('\u0130\u015fveren- Hakedi\u015f Kapsam')) or
            _safe_str(record.get('\u0130\u015fveren- Hakedi\u015f\nKapsam')) or
            _safe_str(record.get('\u0130\u015fveren-Hakedi\u015f Kapsam')) or
            _safe_str(record.get('\u0130\u015fveren-Hakedi\u015f\nKapsam')))


def _get_exchange_rate(work_date):
    """
    Find USD/TRY rate from Döviz Kurları Excel using next-valid-date logic:
    returns the rate for the earliest row whose Tarih >= work_date.
    Returns float or None if the Excel is not uploaded / no matching row.
    """
    rows = _read_excel_rows('doviz_kurlari')
    if not rows:
        return None

    # Detect the date column and the USD/TRY rate column
    date_col = rate_col = None
    for k in rows[0].keys():
        kn = re.sub(r'[^a-z0-9]', '', k.lower())
        if date_col is None and ('tarih' in kn or 'week' in kn or 'month' in kn):
            date_col = k
        if rate_col is None and ('usdtry' in kn or 'usdtl' in kn or 'usd' in kn or 'kur' in kn):
            rate_col = k

    if not date_col or not rate_col:
        return None

    candidates = []
    for row in rows:
        raw_d = row.get(date_col)
        raw_r = row.get(rate_col)
        if raw_d is None or raw_r is None:
            continue
        # Parse date
        d = None
        if hasattr(raw_d, 'date'):           # pandas Timestamp
            d = raw_d.date()
        elif isinstance(raw_d, date):
            d = raw_d
        else:
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d.%m.%Y', '%m/%d/%Y'):
                try:
                    d = datetime.strptime(str(raw_d)[:10], fmt).date()
                    break
                except ValueError:
                    continue
        if d is None:
            continue
        try:
            r = float(str(raw_r).replace(',', '.'))
        except (ValueError, TypeError):
            continue
        if d >= work_date:
            candidates.append((d, r))

    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0])
    return candidates[0][1]


# Employer keyword → contract substring and North/South region
_EMPLOYER_KEYWORD_MAP = {
    'AP-CB':   {'contract_kw': 'APC', 'region': 'South'},
    'BALTIC':  {'contract_kw': 'BAL', 'region': 'North'},
}


def _employer_meta(employer_str):
    """
    Return (contract_keyword, region) for a given employer string.
    Matches by checking if any key in _EMPLOYER_KEYWORD_MAP appears in employer_str.
    Returns ('', None) if no match.
    """
    upper = (employer_str or '').upper()
    for key, meta in _EMPLOYER_KEYWORD_MAP.items():
        if key.upper() in upper:
            return meta['contract_kw'], meta['region']
    return '', None


def _get_rate_cols_for_date(year, month):
    """
    Return (rate_column, currency_column) based on work date year/month.
    Column names match the current hourly_rates Excel headers:
      2023, 2024         → "23-24_Base_Rate_1"       / "23-24 Cur_1"
      2025 month ≤ 6     → "25_ilk6_Base_Rate_2"     / "25_ilk6_Cur_2"
      2025 month >  6     → "25_ikinci6_Base_Rate_3"  / "25_ikinci6_Cur_3"
      2026+              → "26_ilk6_Base_Rate_4"      / "26_ilk6_Cur_4"
    """
    if year in (2023, 2024):
        return "23-24_Base_Rate_1", "23-24 Cur_1"
    elif year == 2025 and month <= 6:
        return "25_ilk6_Base_Rate_2", "25_ilk6_Cur_2"
    elif year == 2025 and month > 6:
        return "25_ikinci6_Base_Rate_3", "25_ikinci6_Cur_3"
    elif year >= 2026:
        return "26_ilk6_Base_Rate_4", "26_ilk6_Cur_4"
    return None, None


def _get_hourly_rate(person_name, work_date, person_id=None):
    """
    Look up the hourly base rate for a person from the Hourly Rates Excel.
    Matches record year/month to the exact VBA column names via _get_rate_cols_for_date.
    Lookup order: ID match first, then name match (contains check).
    Returns (rate: float, currency: str) or (None, None).
    """
    rows = _read_excel_rows('hourly_rates')
    if not rows:
        return None, None

    rate_col, curr_col = _get_rate_cols_for_date(work_date.year, work_date.month)
    if not rate_col:
        return None, None

    # Detect name and ID columns
    name_col = id_col = None
    for k in rows[0].keys():
        kl = k.lower().strip()
        if id_col is None and kl in ('id', 'no', 'no.', 'sicil', 'employee id', 'emp id'):
            id_col = k
        if name_col is None and any(x in kl for x in ('namesur', 'name', 'surname', 'isim', 'personel', 'ad soyad')):
            name_col = k

    # Find matching person row: ID first, then name
    person_row = None
    if person_id is not None and id_col:
        pid_str = str(person_id).strip()
        for row in rows:
            if str(row.get(id_col, '')).strip() == pid_str:
                person_row = row
                break

    if person_row is None and name_col:
        person_lower = (person_name or '').lower().strip()
        if person_lower:
            for row in rows:
                nv = _safe_str(row.get(name_col, '')).lower().strip()
                if nv and (nv == person_lower or nv in person_lower or person_lower in nv):
                    person_row = row
                    break

    if not person_row:
        return None, None

    rate_val = _safe_float(person_row.get(rate_col))
    curr_val = _safe_str(person_row.get(curr_col, 'USD')).upper().strip() if curr_col else 'USD'
    if not curr_val:
        curr_val = 'USD'
    return rate_val, curr_val



def _hakedis_from_record(r, eff_rate):
    """Compute hakedis value from DB record fields, applying eff_rate if needed."""
    rec_isv_currency = _safe_str(
        r.get('İşveren - Currency') or r.get('İşveren- Currency', 'USD')
    ).upper()
    if rec_isv_currency in ('TL', 'TRY'):
        return _safe_float(r.get('İşveren- Hakediş', r.get('İşveren- Hakediş (USD)', 0)))
    else:
        hak_raw = _safe_float(r.get('İşveren- Hakediş (USD)', r.get('İşveren- Hakediş', 0)))
        return hak_raw * eff_rate if eff_rate > 0 else hak_raw


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@hakedis_bp.route('/companies', methods=['GET'])
@role_required('admin', 'project_manager', 'team_leader')
def get_companies(current_user=None):
    """Return distinct Company values."""
    records = _load_records()
    companies = sorted({_safe_str(r.get('Company')) for r in records if _safe_str(r.get('Company'))})
    return jsonify({'success': True, 'data': companies})


@hakedis_bp.route('/contracts', methods=['GET'])
@role_required('admin', 'project_manager', 'team_leader')
def get_contracts(current_user=None):
    """Return distinct İşveren- Sözleşme No values for the company.
    Also returns Scope and Projects/Group values as fallback groupings.
    """
    company = request.args.get('company', '').strip()
    records = _load_records()

    contracts = set()
    scopes    = set()
    projects  = set()
    for r in records:
        if company and _safe_str(r.get('Company')) != company:
            continue
        c = _safe_str(r.get('İşveren- Sözleşme No', ''))
        if c:
            contracts.add(c)
        s = _safe_str(r.get('Scope', ''))
        if s:
            scopes.add(s)
        p = _safe_str(r.get('Projects/Group', '') or r.get('Projects', ''))
        if p:
            projects.add(p)

    return jsonify({
        'success':   True,
        'data':      sorted(contracts),
        'scopes':    sorted(scopes),
        'projects':  sorted(projects),
        'hasContracts': len(contracts) > 0,
    })


@hakedis_bp.route('/generate', methods=['POST'])
@role_required('admin', 'project_manager', 'team_leader')
def generate_hakedis(current_user=None):
    """
    Generate a Hakedis (Progress Payment) report.

    Expected JSON body:
    {
        "company":     "AP-CB Pergel",
        "contractNo":  "RHI-01856",
        "startDate":   "2024-12-01",   // period start (YYYY-MM-DD)
        "endDate":     "2024-12-31",   // period end
        "reportDate":  "2025-01-15",   // optional
        "periodText":  "Aralık 2024",  // optional, e.g. "PP-11"
        "format":      "excel"         // "excel" or "json" (default: "json")
    }
    """
    body = request.get_json(force=True) or {}

    company      = body.get('company', '').strip()
    employer     = body.get('employer', '').strip()      # İşveren / client name
    contract_no  = body.get('contractNo', '').strip()   # optional
    scope_filter = body.get('scope', '').strip()        # optional
    usd_to_tl    = float(body.get('usdToTlRate', 0) or 0)  # 0 = keep USD
    start_str    = body.get('startDate', '')
    end_str      = body.get('endDate', '')
    report_date_str = body.get('reportDate', '')
    period_text  = body.get('periodText', '')
    fmt          = body.get('format', 'json')
    # Currency label for output
    output_currency = 'TL' if usd_to_tl > 0 else 'USD'
    # (May be overridden to 'TL' below if Döviz Kurları Excel auto-rate is used)

    try:
        start_date = datetime.strptime(start_str[:10], '%Y-%m-%d').date()
        end_date   = datetime.strptime(end_str[:10],   '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400

    report_date = date.today()
    if report_date_str:
        try:
            report_date = datetime.strptime(report_date_str[:10], '%Y-%m-%d').date()
        except ValueError:
            pass

    # Cumulative = Jan 1 of startDate year  →  day before startDate
    year_start  = date(start_date.year, 1, 1)
    day_before  = start_date - timedelta(days=1)

    # Employer → North/South regional routing
    _, ns_region = _employer_meta(employer)

    # Determine if auto exchange rate is available (Döviz Kurları Excel uploaded)
    use_auto_rate = bool(_read_excel_rows('doviz_kurlari'))
    if use_auto_rate:
        output_currency = 'TL'   # Auto rates always convert to TL
    # Determine if Hourly Rates Excel is available
    use_hourly_rates = bool(_read_excel_rows('hourly_rates'))

    # --- Load company name mapping (Şirket İsimleri) ---
    company_name_map = {}
    sirket_rows = _read_excel_rows('sirket_isimleri')
    if sirket_rows and len(sirket_rows) > 0:
        # Try various column name combinations for short/long names
        for row in sirket_rows:
            short_name = (_safe_str(row.get('Şirket')) or
                         _safe_str(row.get('Sirket')) or
                         _safe_str(row.get('Kısa Ad')) or 
                         _safe_str(row.get('Short Name')) or
                         _safe_str(row.get('Kisa Ad')) or
                         _safe_str(row.get('ShortName')) or
                         _safe_str(row.get('short')) or
                         _safe_str(row.get('Short')) or
                         _safe_str(row.get('Kısa İsim')) or
                         _safe_str(row.get('Kisa Isim')))
            long_name = (_safe_str(row.get('Şirket Tam İsim')) or
                        _safe_str(row.get('Sirket Tam Isim')) or
                        _safe_str(row.get('Tam İsim')) or
                        _safe_str(row.get('Tam Isim')) or
                        _safe_str(row.get('Uzun Ad')) or
                        _safe_str(row.get('Long Name')) or
                        _safe_str(row.get('LongName')) or
                        _safe_str(row.get('long')) or
                        _safe_str(row.get('Long')) or
                        _safe_str(row.get('Full Name')) or
                        _safe_str(row.get('FullName')) or
                        _safe_str(row.get('Uzun İsim')) or
                        _safe_str(row.get('Uzun Isim')))
            if short_name and long_name:
                company_name_map[short_name] = long_name
    
    # Map company and employer to long names if available
    company_long = company_name_map.get(company, company)
    employer_long = company_name_map.get(employer, employer)

    # --- Load & filter records ---
    # Records are matched by Company + date range (+ optional N/S region + scope).
    # Sözleşme No from the Hakedis Excel is a report label only — not stored in DB records.
    all_records = _load_records()
    matching = []
    for r in all_records:
        if _safe_str(r.get('Company')) != company:
            continue
        # Regional filter: AP-CB → South, BALTIC → North
        if ns_region:
            rec_region = _safe_str(r.get('North/\nSouth', '') or r.get('North/ South', '') or r.get('North/South', '')).strip()
            if rec_region and rec_region != ns_region:
                continue
        # Optional scope filter
        if scope_filter:
            if _safe_str(r.get('Scope', '')) != scope_filter:
                continue
        matching.append(r)

    if not matching:
        return jsonify({
            'success': True,
            'data': [],
            'meta': {
                'company': company,
                'employer': employer,
                'contractNo': contract_no,
                'startDate': start_str,
                'endDate': end_str,
                'totalRecords': 0,
            }
        })

    # --- Aggregate (VBA-matching: row-by-row formula, grouped by Projects) ---
    # VBA: SatirTutar = Saat * NihaiBirimFiyat  where
    #       NihaiBirimFiyat = HamUcret * AnlikDolarKuru   (if USD)
    #       NihaiBirimFiyat = HamUcret                    (if TL)
    # Accumulate SatirTutar per rs![Projects]

    current_sum  = defaultdict(float)   # project → TL/currency amount (current period)
    current_mh   = defaultdict(float)   # project → raw MH (current period)
    current_pers = defaultdict(set)     # project → set of persons
    previous_sum = defaultdict(float)   # project → amount (year-to-date before start)
    previous_mh  = defaultdict(float)   # project → raw MH (previous period)
    pp_months    = defaultdict(set)     # project → set of (year, month)
    disciplines  = defaultdict(set)
    scopes       = defaultdict(set)
    # Per-person detail: {project: {person: {'cur_mh','cur_total','prev_mh','prev_total','rate','currency'}}}
    person_detail = defaultdict(lambda: defaultdict(lambda: {
        'cur_mh': 0.0, 'cur_total': 0.0,
        'prev_mh': 0.0, 'prev_total': 0.0,
        'rate': None, 'currency': ''
    }))

    for r in matching:
        rec_date = _parse_record_date(r)
        if rec_date is None:
            continue

        # VBA groups by rs![Projects]
        project = (_safe_str(r.get('Projects')) or
                   _safe_str(r.get('Projects/Group')) or
                   'Unknown')

        # MH: VBA tries TOTAL_MH first, then TOTAL MH
        mh = (_safe_float(r.get('TOTAL_MH')) or _get_mh(r))

        disc   = _safe_str(r.get('Discipline'))
        scope  = _get_kapsam(r) or _safe_str(r.get('Scope'))
        person = _safe_str(r.get('Name Surname'))
        person_id = r.get('ID')

        # --- VBA row-by-row formula ---
        if use_hourly_rates:
            h_rate, h_curr = _get_hourly_rate(person, rec_date, person_id)
            if h_rate and h_rate > 0:
                if h_curr in ('TL', 'TRY'):
                    # TL rate: no currency conversion
                    nihai_birim = h_rate
                else:
                    # USD rate: multiply by exchange rate for this record's date
                    kur = _get_exchange_rate(rec_date) or usd_to_tl or 1.0
                    nihai_birim = h_rate * kur
                satir_tutar = mh * nihai_birim
            else:
                satir_tutar = 0.0
        else:
            # No Hourly Rates file: use raw exchange rate × MH
            eff = (_get_exchange_rate(rec_date) if use_auto_rate else usd_to_tl) or 1.0
            satir_tutar = mh * eff

        # PP month tracking (for whole year including current)
        if year_start <= rec_date <= end_date:
            pp_months[project].add((rec_date.year, rec_date.month))

        # Current period
        if start_date <= rec_date <= end_date:
            current_sum[project]  += satir_tutar
            current_mh[project]   += mh
            current_pers[project].add(person)
            if disc:
                disciplines[project].add(disc)
            if scope:
                scopes[project].add(scope)
            if person:
                pd = person_detail[project][person]
                pd['cur_mh']    += mh
                pd['cur_total'] += satir_tutar
                if use_hourly_rates and h_rate:
                    pd['rate']     = h_rate
                    pd['currency'] = h_curr

        # Previous period (same year, before start_date)
        elif year_start <= rec_date <= day_before:
            previous_sum[project] += satir_tutar
            previous_mh[project]  += mh
            if person:
                pd = person_detail[project][person]
                pd['prev_mh']    += mh
                pd['prev_total'] += satir_tutar
                if use_hourly_rates and h_rate and pd['rate'] is None:
                    pd['rate']     = h_rate
                    pd['currency'] = h_curr

    # Collect all projects that appear in either period
    all_projects = sorted(set(current_sum.keys()) | set(previous_sum.keys()))

    rows = []
    for proj in all_projects:
        pp_no = len(pp_months.get(proj, set()))
        cur_hakedis  = current_sum[proj]
        prev_hakedis = previous_sum[proj]
        # Cumulative = Geçmiş Dönem + Cari Dönem
        cumulative_hakedis = prev_hakedis + cur_hakedis

        rows.append({
            'project':           proj,
            'ppNo':              pp_no,
            'currentMH':         round(current_mh[proj], 2),
            'currentHakedis':    round(cur_hakedis, 2),
            'previousMH':        round(previous_mh[proj], 2),
            'previousHakedis':   round(prev_hakedis, 2),
            'cumulativeHakedis': round(cumulative_hakedis, 2),
            'disciplines':       sorted(disciplines.get(proj, set())),
            'scope':             sorted(scopes.get(proj, set())),
            'personCount':       len(current_pers.get(proj, set())),
        })


    # Sort by project name
    rows.sort(key=lambda x: x['project'])

    # Totals row
    totals = {
        'project':          'TOPLAM',
        'ppNo':             '',
        'currentMH':        round(sum(r['currentMH']        for r in rows), 2),
        'currentHakedis':   round(sum(r['currentHakedis']   for r in rows), 2),
        'previousMH':       round(sum(r['previousMH']       for r in rows), 2),
        'previousHakedis':  round(sum(r['previousHakedis']  for r in rows), 2),
        'cumulativeHakedis':round(sum(r['cumulativeHakedis']for r in rows), 2),
        'disciplines':      [],
        'scope':            [],
        'personCount':      '',
    }

    meta = {
        'company':         company_long,
        'employer':        employer_long,
        'contractNo':      contract_no,
        'startDate':       start_str,
        'endDate':         end_str,
        'reportDate':      report_date.strftime('%Y-%m-%d'),
        'periodText':      period_text,
        'totalRecords':    len(matching),
        'currentRecords':  sum(
            1 for r in matching
            if (d := _parse_record_date(r)) and start_date <= d <= end_date
        ),
        'currency':        output_currency,
        'usdToTlRate':     usd_to_tl if usd_to_tl > 0 else None,
        'autoRate':        use_auto_rate,
        'hourlyRates':     use_hourly_rates,
        'nsRegion':        ns_region,
    }

    if fmt == 'excel':
        excel_bytes = _build_excel(rows, totals, meta, dict(person_detail))
        buf = io.BytesIO(excel_bytes)
        buf.seek(0)
        filename = f"Hakedis_{company.replace(' ', '_')}_{contract_no}_{start_str[:7]}.xlsx"
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    return jsonify({'success': True, 'data': rows, 'totals': totals, 'meta': meta})


# ---------------------------------------------------------------------------
# Bulk generate — all companies in a period → ZIP with two folders
# ---------------------------------------------------------------------------

@hakedis_bp.route('/bulk-generate', methods=['POST'])
@role_required('admin', 'project_manager', 'team_leader')
def bulk_generate_hakedis(current_user=None):
    """
    Generate progress payment Excel files for ALL companies that have data in a
    given date range, then return a single ZIP file containing two sub-folders:
      - "AP-CB Hakediş/"  (South region / APC contracts)
      - "BALTIC Hakediş/" (North region / BAL contracts)
    Each file is named  <Company>_<YYYY-MM>.xlsx.

    Expected JSON body:
    {
        "startDate":   "2025-12-01",
        "endDate":     "2025-12-31",
        "reportDate":  "2026-01-01",   // optional
        "periodText":  "Aralık 2025",  // optional
        "usdToTlRate": 38              // ignored when Döviz Kurları Excel is uploaded
    }
    """
    import zipfile

    body = request.get_json(force=True) or {}
    start_str       = body.get('startDate', '')
    end_str         = body.get('endDate', '')
    period_text     = body.get('periodText', '')
    usd_to_tl       = float(body.get('usdToTlRate', 0) or 0)
    report_date_str = body.get('reportDate', '')

    try:
        start_date = datetime.strptime(start_str[:10], '%Y-%m-%d').date()
        end_date   = datetime.strptime(end_str[:10],   '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'error': 'Geçersiz tarih formatı. YYYY-MM-DD kullanın.'}), 400

    report_date = date.today()
    if report_date_str:
        try:
            report_date = datetime.strptime(report_date_str[:10], '%Y-%m-%d').date()
        except ValueError:
            pass

    year_start = date(start_date.year, 1, 1)
    day_before = start_date - timedelta(days=1)

    use_auto_rate    = bool(_read_excel_rows('doviz_kurlari'))
    use_hourly_rates = bool(_read_excel_rows('hourly_rates'))
    if use_auto_rate:
        usd_to_tl = 0
    output_currency = 'TL' if (usd_to_tl > 0 or use_auto_rate) else 'USD'

    # Company name mapping (short → long)
    company_name_map = {}
    for row in (_read_excel_rows('sirket_isimleri') or []):
        short_name = (_safe_str(row.get('Şirket')) or _safe_str(row.get('Sirket')) or
                      _safe_str(row.get('Kısa Ad')) or _safe_str(row.get('Short Name')) or
                      _safe_str(row.get('Kisa Ad')))
        long_name  = (_safe_str(row.get('Şirket Tam İsim')) or _safe_str(row.get('Sirket Tam Isim')) or
                      _safe_str(row.get('Tam İsim')) or _safe_str(row.get('Uzun Ad')) or
                      _safe_str(row.get('Long Name')))
        if short_name and long_name:
            company_name_map[short_name] = long_name

    # Contracts from Hakedis Excel, keyed by company name and filtered by period year
    # Logic mirrors the cascade endpoint: 2026 also accepts 2025 contracts
    _BULK_TR = str.maketrans('\u00e7\u011f\u0131\u00f6\u015f\u00fc\u00c7\u011e\u0130\u00d6\u015e\u00dc', 'cgiosucgiosu')
    def _bn(s):
        return re.sub(r'[^a-z0-9]', '', (s or '').lower().translate(_BULK_TR))

    hakedis_contracts_map = {}   # { firm_name: [contract_no, ...] }
    CONTRACT_RE2 = re.compile(r'20\d{2}[-/]\d{2}[-/]\d{2}')
    hakedis_excel_rows = _read_excel_rows('hakedis')
    if hakedis_excel_rows:
        sozlesme_col = firma_col = None
        for k in hakedis_excel_rows[0].keys():
            kn = _bn(k)
            if sozlesme_col is None and 'sozlesme' in kn:
                sozlesme_col = k
            if firma_col is None and ('firma' in kn or 'company' in kn or 'sirket' in kn):
                firma_col = k
        # Fallback: detect by value pattern
        if sozlesme_col is None:
            for k in hakedis_excel_rows[0].keys():
                hits = sum(1 for row in hakedis_excel_rows[:10]
                           if CONTRACT_RE2.match(str(row.get(k) or '').strip()))
                if hits >= max(1, min(len(hakedis_excel_rows[:10]), 3)):
                    sozlesme_col = k
                    break
        if sozlesme_col:
            year_int = start_date.year
            for row in hakedis_excel_rows:
                con = str(row.get(sozlesme_col, '') or '').strip()
                if not con or con in ('nan', 'None'):
                    continue
                # Year filter: 2026 period also accepts 2025 contracts
                if year_int == 2026:
                    if not (con.startswith('2026') or con.startswith('2025')):
                        continue
                else:
                    if not con.startswith(str(year_int)):
                        continue
                firm = str(row.get(firma_col, '') or '').strip() if firma_col else ''
                if firm in ('nan', 'None'):
                    firm = ''
                hakedis_contracts_map.setdefault(firm, set()).add(con)
        hakedis_contracts_map = {k: sorted(v) for k, v in hakedis_contracts_map.items()}

    all_records = _load_records()

    # ── Pass 1: find (company, employer_key) combos with current-period data ──
    combo_set = set()   # {(company, 'AP-CB') | (company, 'BALTIC')}
    for r in all_records:
        rec_date = _parse_record_date(r)
        if not rec_date or not (start_date <= rec_date <= end_date):
            continue
        comp = _safe_str(r.get('Company', ''))
        if not comp:
            continue
        # Skip companies that are themselves employers (AP-CB, BALTIC, etc.)
        if comp.strip().upper() in {k.upper() for k in _EMPLOYER_KEYWORD_MAP}:
            continue
        ns       = _safe_str(r.get('North/\nSouth', '') or r.get('North/ South', '') or r.get('North/South', '')).strip()
        contract = _safe_str(r.get('İşveren- Sözleşme No', '')).upper()
        if ns == 'South' or 'APC' in contract:
            combo_set.add((comp, 'AP-CB'))
        elif ns == 'North' or 'BAL' in contract:
            combo_set.add((comp, 'BALTIC'))
        else:
            # Fallback: assign to both employers if region can't be determined
            combo_set.add((comp, 'AP-CB'))
            combo_set.add((comp, 'BALTIC'))

    if not combo_set:
        return jsonify({'success': False, 'error': 'Seçilen dönemde hiçbir şirkete ait veri bulunamadı.'}), 404

    # ── Pass 2: aggregate + build Excel per (company, employer) ──
    zip_buf   = io.BytesIO()
    generated = []
    errors    = []

    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for employer_key in ('AP-CB', 'BALTIC'):
            _, ns_region = _employer_meta(employer_key)
            folder_name  = f'{employer_key} Hakediş'

            for comp in sorted(c for (c, e) in combo_set if e == employer_key):
                try:
                    # Collect all year-to-date records for this company + region
                    matching = []
                    for r in all_records:
                        if _safe_str(r.get('Company')) != comp:
                            continue
                        rec_ns = _safe_str(
                            r.get('North/\nSouth', '') or r.get('North/ South', '') or r.get('North/South', '')
                        ).strip()
                        if ns_region and rec_ns and rec_ns != ns_region:
                            continue
                        rec_date = _parse_record_date(r)
                        if rec_date and year_start <= rec_date <= end_date:
                            matching.append(r)

                    if not matching:
                        continue

                    # Aggregation (mirrors generate_hakedis logic exactly)
                    current_sum  = defaultdict(float)
                    current_mh   = defaultdict(float)
                    current_pers = defaultdict(set)
                    previous_sum = defaultdict(float)
                    previous_mh  = defaultdict(float)
                    pp_months    = defaultdict(set)
                    disciplines  = defaultdict(set)
                    scopes_d     = defaultdict(set)
                    person_detail = defaultdict(lambda: defaultdict(lambda: {
                        'cur_mh': 0.0, 'cur_total': 0.0,
                        'prev_mh': 0.0, 'prev_total': 0.0,
                        'rate': None, 'currency': '',
                    }))

                    for r in matching:
                        rec_date = _parse_record_date(r)
                        if not rec_date:
                            continue
                        project  = (_safe_str(r.get('Projects')) or
                                    _safe_str(r.get('Projects/Group')) or 'Unknown')
                        mh       = _safe_float(r.get('TOTAL_MH')) or _get_mh(r)
                        disc     = _safe_str(r.get('Discipline'))
                        scope    = _get_kapsam(r) or _safe_str(r.get('Scope'))
                        person   = _safe_str(r.get('Name Surname'))
                        pid      = r.get('ID')

                        if use_hourly_rates:
                            h_rate, h_curr = _get_hourly_rate(person, rec_date, pid)
                            if h_rate and h_rate > 0:
                                if h_curr in ('TL', 'TRY'):
                                    nihai = h_rate
                                else:
                                    kur   = _get_exchange_rate(rec_date) or usd_to_tl or 1.0
                                    nihai = h_rate * kur
                                satir = mh * nihai
                            else:
                                satir = 0.0
                        else:
                            eff   = (_get_exchange_rate(rec_date) if use_auto_rate else usd_to_tl) or 1.0
                            satir = mh * eff

                        if year_start <= rec_date <= end_date:
                            pp_months[project].add((rec_date.year, rec_date.month))

                        if start_date <= rec_date <= end_date:
                            current_sum[project]  += satir
                            current_mh[project]   += mh
                            current_pers[project].add(person)
                            if disc:   disciplines[project].add(disc)
                            if scope:  scopes_d[project].add(scope)
                            if person:
                                pd = person_detail[project][person]
                                pd['cur_mh']    += mh
                                pd['cur_total'] += satir
                                if use_hourly_rates and h_rate:
                                    pd['rate']     = h_rate
                                    pd['currency'] = h_curr
                        elif year_start <= rec_date <= day_before:
                            previous_sum[project] += satir
                            previous_mh[project]  += mh
                            if person:
                                pd = person_detail[project][person]
                                pd['prev_mh']    += mh
                                pd['prev_total'] += satir
                                if use_hourly_rates and h_rate and pd['rate'] is None:
                                    pd['rate']     = h_rate
                                    pd['currency'] = h_curr

                    all_projects = sorted(set(current_sum.keys()) | set(previous_sum.keys()))
                    if not all_projects:
                        continue

                    rows_out = []
                    for proj in all_projects:
                        ch = current_sum[proj]
                        ph = previous_sum[proj]
                        rows_out.append({
                            'project':           proj,
                            'ppNo':              len(pp_months.get(proj, set())),
                            'currentMH':         round(current_mh[proj], 2),
                            'currentHakedis':    round(ch, 2),
                            'previousMH':        round(previous_mh[proj], 2),
                            'previousHakedis':   round(ph, 2),
                            'cumulativeHakedis': round(ch + ph, 2),
                            'disciplines':       sorted(disciplines.get(proj, set())),
                            'scope':             sorted(scopes_d.get(proj, set())),
                            'personCount':       len(current_pers.get(proj, set())),
                        })
                    rows_out.sort(key=lambda x: x['project'])

                    totals_out = {
                        'project':           'TOPLAM', 'ppNo': '',
                        'currentMH':         round(sum(x['currentMH']         for x in rows_out), 2),
                        'currentHakedis':    round(sum(x['currentHakedis']    for x in rows_out), 2),
                        'previousMH':        round(sum(x['previousMH']        for x in rows_out), 2),
                        'previousHakedis':   round(sum(x['previousHakedis']   for x in rows_out), 2),
                        'cumulativeHakedis': round(sum(x['cumulativeHakedis'] for x in rows_out), 2),
                        'disciplines': [], 'scope': [], 'personCount': '',
                    }

                    # Look up contract(s) from Hakedis Excel for this company + year;
                    # try exact match → case-insensitive → normalized partial → broadcast (no FIRMA col) → DB fallback
                    contract_nos = (
                        hakedis_contracts_map.get(comp)
                        or next((v for k, v in hakedis_contracts_map.items()
                                 if k and k.lower() == comp.lower()), None)
                        or next((v for k, v in hakedis_contracts_map.items()
                                 if k and (_bn(k) in _bn(comp) or _bn(comp) in _bn(k))), None)
                        or hakedis_contracts_map.get('')   # broadcast when no FIRMA column
                        or sorted({
                            _safe_str(r.get('İşveren- Sözleşme No', ''))
                            for r in matching
                            if _safe_str(r.get('İşveren- Sözleşme No', ''))
                        })
                    ) or []

                    # Filter by employer keyword: AP-CB → keep only *-APC-* contracts,
                    # BALTIC → keep only *-BAL-* contracts
                    emp_kw = (_EMPLOYER_KEYWORD_MAP.get(employer_key) or {}).get('contract_kw', '')
                    if emp_kw and contract_nos:
                        filtered = [c for c in contract_nos if emp_kw.upper() in c.upper()]
                        if filtered:  # only apply filter if it leaves at least one result
                            contract_nos = filtered

                    # Prefer addendum contracts (-Z01, -Z02, etc.) over base contracts.
                    # If any contract ends with -Z followed by digits, keep only those.
                    addendum = [c for c in contract_nos if re.search(r'-Z\d+$', c, re.IGNORECASE)]
                    if addendum:
                        contract_nos = addendum

                    meta_out = {
                        'company':        company_name_map.get(comp, comp),
                        'employer':       company_name_map.get(employer_key, employer_key),
                        'contractNo':     ', '.join(contract_nos),
                        'startDate':      start_str,
                        'endDate':        end_str,
                        'reportDate':     report_date.strftime('%Y-%m-%d'),
                        'periodText':     period_text,
                        'totalRecords':   len(matching),
                        'currentRecords': sum(
                            1 for r in matching
                            if (d := _parse_record_date(r)) and start_date <= d <= end_date
                        ),
                        'currency':       output_currency,
                        'usdToTlRate':    usd_to_tl if usd_to_tl > 0 else None,
                        'autoRate':       use_auto_rate,
                        'hourlyRates':    use_hourly_rates,
                        'nsRegion':       ns_region,
                    }

                    excel_bytes  = _build_excel(rows_out, totals_out, meta_out, dict(person_detail))
                    safe_name    = comp.replace(' ', '_').replace('/', '-')
                    zip_path     = f'{folder_name}/{safe_name}_{start_str[:7]}.xlsx'
                    zf.writestr(zip_path, excel_bytes)
                    generated.append({'company': comp, 'employer': employer_key})

                except Exception as exc:
                    errors.append(f'{comp} ({employer_key}): {exc}')

    if not generated:
        msg = 'Hiçbir Excel oluşturulamadı.'
        if errors:
            msg += ' Hatalar: ' + '; '.join(errors)
        return jsonify({'success': False, 'error': msg}), 500

    zip_buf.seek(0)
    return send_file(
        zip_buf,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'Toplu_Hakedis_{start_str[:7]}.zip',
    )


# ---------------------------------------------------------------------------
# Multi-Excel upload / status / delete endpoints
# ---------------------------------------------------------------------------

def _read_excel_rows(excel_type):
    """
    Read the stored Excel file for the given type.
    Auto-detects which sheet to read (tries the type-specific sheet name first,
    then Sheet1, then the first sheet).
    Returns a list of dicts (column → value).
    """
    try:
        import pandas as pd
    except ImportError:
        return []

    info = _excel_store.get(excel_type)
    if not info or not os.path.exists(info['path']):
        return []

    try:
        xl = pd.ExcelFile(info['path'])
        sheet_labels = {
            'database':        ['DATABASE', 'Database'],
            'hourly_rates':    ['Hourly Rates', 'HourlyRates', 'hourly rates'],
            'hakedis':         ['Hakedis', 'Hakediş', 'hakedis'],
            'doviz_kurlari':   ['Doviz Kurlari', 'Döviz Kurları', 'DovizKurlari'],
            'sirket_isimleri': ['tbl_SirketIsimleri', 'SirketIsimleri', 'Sirket'],
            'info':            ['Info', 'INFO'],
        }
        preferred = sheet_labels.get(excel_type, [])
        sheet_name = None
        for p in preferred:
            if p in xl.sheet_names:
                sheet_name = p
                break
        if sheet_name is None:
            sheet_name = xl.sheet_names[0]

        # Hourly Rates Excel: auto-detect single-row vs two-row merged header.
        #   OLD format: Row 0 = year group labels (e.g. "2023-2024", "2025 İlk 6 ay")
        #               Row 1 = sub-headers ("Currency 1", "Hourly Base Rates 1", …)
        #   NEW format: Row 0 = actual column names (e.g. "23-24 Cur_1", "26_ilk6_Base_Rate_4")
        if excel_type == 'hourly_rates':
            # Peek at row 0 to decide
            df_peek = pd.read_excel(info['path'], sheet_name=sheet_name, header=0, nrows=0, dtype=str)
            peek_cols = [str(c).strip() for c in df_peek.columns]
            # New format: columns already contain rate/currency keywords
            _new_fmt_kw = ('_cur_', '_base_rate_', 'add_cur', 'add_rate', 'cur_1', 'base_rate_1')
            is_new_fmt = any(any(kw in c.lower() for kw in _new_fmt_kw) for c in peek_cols)
            if is_new_fmt:
                # Single-row header — read directly
                df = pd.read_excel(info['path'], sheet_name=sheet_name, header=0, dtype=str)
            else:
                # Old two-row merged header — MultiIndex flatten
                df = pd.read_excel(info['path'], sheet_name=sheet_name, header=[0, 1], dtype=str)
                flat_cols = []
                for parent, child in df.columns:
                    p = str(parent).strip()
                    c = str(child).strip()
                    if p.startswith('Unnamed') or p == '':
                        flat_cols.append(c)
                    else:
                        flat_cols.append(f'{p} {c}')
                df.columns = flat_cols
            df = df.fillna('')
            # Drop rows that are entirely empty
            df = df[df.apply(lambda r: any(v.strip() for v in r.values if isinstance(v, str)), axis=1)]
            return df.to_dict('records')

        df = pd.read_excel(info['path'], sheet_name=sheet_name, dtype=str)
        df = df.fillna('')
        return df.to_dict('records')
    except Exception as e:
        print(f'[EXCEL READ] {excel_type}: {e}')
        return []


@hakedis_bp.route('/excel-status', methods=['GET'])
@role_required('admin', 'project_manager', 'team_leader')
def excel_status(current_user=None):
    """Return upload status for all 6 Excel slots."""
    result = {}
    for key, meta in EXCEL_TYPES.items():
        stored = _excel_store.get(key)
        result[key] = {
            'label':       meta['label'],
            'description': meta['description'],
            'uploaded':    bool(stored),
            'filename':    stored['filename'] if stored else None,
            'uploadedAt':  stored['uploaded_at'] if stored else None,
        }
    return jsonify({'success': True, 'data': result})


@hakedis_bp.route('/preview-excel/<excel_type>', methods=['GET'])
@role_required('admin', 'project_manager', 'team_leader')
def preview_excel(excel_type, current_user=None):
    """Return up to 200 rows from a stored Excel file for preview."""
    if excel_type not in EXCEL_TYPES:
        return jsonify({'success': False, 'error': f'Unknown type "{excel_type}"'}), 400
    stored = _excel_store.get(excel_type)
    if not stored:
        return jsonify({'success': False, 'error': 'Not uploaded yet'}), 404
    try:
        import pandas as pd
        path = stored['path']
        xl   = pd.ExcelFile(path)
        df   = pd.read_excel(path, sheet_name=xl.sheet_names[0], nrows=200, dtype=str)
        df   = df.fillna('')
        columns = list(df.columns)
        rows    = df.values.tolist()
        return jsonify({'success': True, 'columns': columns, 'rows': rows, 'total': len(rows)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@hakedis_bp.route('/upload-excel', methods=['POST'])
@role_required('admin', 'project_manager', 'team_leader')
def upload_excel_file(current_user=None):
    """
    Upload one of the 6 Excel files.
    Form fields: file (xlsx/xls), type (one of EXCEL_TYPES keys).
    """
    excel_type = request.form.get('type', '').strip()
    if excel_type not in EXCEL_TYPES:
        return jsonify({'success': False, 'error': f'Unknown type "{excel_type}". Valid: {list(EXCEL_TYPES.keys())}'}), 400

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file in request'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in {'.xlsx', '.xls', '.xlsb'}:
        return jsonify({'success': False, 'error': f'Unsupported extension "{ext}". Use .xlsx or .xls'}), 400

    safe_name = secure_filename(f.filename)
    save_path = os.path.join(UPLOAD_DIR, f'hakedis_{excel_type}{ext}')
    f.save(save_path)

    # Read a quick row-count preview
    row_count = 0
    columns   = []
    try:
        import pandas as pd
        xl = pd.ExcelFile(save_path)
        df = pd.read_excel(save_path, sheet_name=xl.sheet_names[0], nrows=3)
        columns   = list(df.columns)
        # Read full sheet just for count
        df_full   = pd.read_excel(save_path, sheet_name=xl.sheet_names[0])
        row_count = len(df_full)
    except Exception as e:
        print(f'[EXCEL UPLOAD] preview failed: {e}')

    _excel_store[excel_type] = {
        'path':        save_path,
        'filename':    f.filename,
        'uploaded_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'row_count':   row_count,
        'columns':     columns[:20],
    }

    return jsonify({
        'success':  True,
        'type':     excel_type,
        'filename': f.filename,
        'rows':     row_count,
        'columns':  columns[:20],
    })


@hakedis_bp.route('/delete-excel/<excel_type>', methods=['DELETE'])
@role_required('admin', 'project_manager', 'team_leader')
def delete_excel_file(excel_type, current_user=None):
    """Delete a previously uploaded Excel file."""
    if excel_type not in EXCEL_TYPES:
        return jsonify({'success': False, 'error': 'Unknown type'}), 400

    stored = _excel_store.pop(excel_type, None)
    if stored and os.path.exists(stored['path']):
        try:
            os.remove(stored['path'])
        except Exception:
            pass

    return jsonify({'success': True, 'deleted': excel_type})


# ---------------------------------------------------------------------------
# Access Database upload + parse endpoints
# ---------------------------------------------------------------------------

@hakedis_bp.route('/upload-access', methods=['POST'])
@role_required('admin', 'project_manager', 'team_leader')
def upload_access(current_user=None):
    """
    Upload a Microsoft Access .accdb / .mdb file.
    Parses:
      - All table names + column info
      - All saved query/view definitions
      - All VBA module code (via oletools)
    Returns a comprehensive structure summary.
    """
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file part in request'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_ACCESS_EXTENSIONS:
        return jsonify({'success': False, 'error': f'Unsupported file type "{ext}". Upload .accdb or .mdb'}), 400

    safe_name = secure_filename(f.filename)
    save_path = os.path.join(UPLOAD_DIR, 'access_' + safe_name)
    f.save(save_path)

    result = _parse_access_file(save_path)
    result['filename'] = f.filename
    result['savedPath'] = save_path

    # Cache globally so later hakedis generation can use it
    _access_store.clear()
    _access_store.update(result)

    return jsonify({'success': True, 'data': result})


@hakedis_bp.route('/access-structure', methods=['GET'])
@role_required('admin', 'project_manager', 'team_leader')
def get_access_structure(current_user=None):
    """Return the last parsed Access structure (if any)."""
    if not _access_store:
        return jsonify({'success': False, 'error': 'No Access file uploaded yet'}), 404
    return jsonify({'success': True, 'data': dict(_access_store)})


@hakedis_bp.route('/access-contracts', methods=['GET'])
@role_required('admin', 'project_manager', 'team_leader')
def get_access_contracts(current_user=None):
    """
    Return contract numbers read directly from the Access Sözleşme table.
    If no Access file has been uploaded, falls back to the PostgreSQL records.
    """
    if _access_store and _access_store.get('contractsFromAccess'):
        rows = _access_store['contractsFromAccess']
        # Try to find a column that looks like a contract number
        contract_nums = set()
        for row in rows:
            for k, v in row.items():
                k_norm = re.sub(r'[^a-z0-9]', '', k.lower())
                if any(x in k_norm for x in ('no', 'num', 'kod', 'code', 'sozlesme', 'contract')):
                    if v and str(v).strip() not in ('None', ''):
                        contract_nums.add(str(v).strip())
        return jsonify({
            'success': True,
            'source':  'access',
            'data':    sorted(contract_nums),
            'allRows': rows,
        })

    # Fallback: PostgreSQL records
    company = request.args.get('company', '').strip()
    records = _load_records()
    seen = set()
    for r in records:
        if company and _safe_str(r.get('Company')) != company:
            continue
        c = _safe_str(r.get('İşveren- Sözleşme No', ''))
        if c:
            seen.add(c)
    return jsonify({'success': True, 'source': 'db', 'data': sorted(seen)})


@hakedis_bp.route('/cascade', methods=['GET'])
@role_required('admin', 'project_manager', 'team_leader')
def get_cascade(current_user=None):
    """
    Return a 3-level cascade structure built from the Hakedis Excel + DB:
      employers  → list of distinct employer names
      contracts  → { employer: [contractNo, ...] }
      scopes     → { contractNo: [scope, ...] }

    Optional query params:
      employer   → apply keyword filtering (AP-CB→APC, BALTIC→BAL)
      year       → filter contracts starting with this year string;
                   year==2026 also includes 2025 contracts
    """
    # Optional filter params
    cascade_employer = request.args.get('employer', '').strip()
    cascade_year     = request.args.get('year', '').strip()    # e.g. '2025', '2026'
    CONTRACT_RE = re.compile(r'20\d{2}[-/]\d{2}[-/]\d{2}')

    _TR_TABLE = str.maketrans('çğıöşüÇĞİÖŞÜ', 'cgiosucgiosu')

    def _norm_col(name):
        s = (name or '').lower().translate(_TR_TABLE)
        return re.sub(r'[^a-z0-9]', '', s)

    def _detect_columns(rows):
        if not rows:
            return None, None, None, None
        sample = rows[0]
        employer_col = contract_col = scope_col = project_col = None
        # Pass 1 – name-based heuristics
        for k in sample.keys():
            kn = _norm_col(k)
            if employer_col is None and any(x in kn for x in
                    ('isveren', 'isvern', 'employer', 'issuer', 'musteri', 'client')):
                employer_col = k
            elif contract_col is None and any(x in kn for x in
                    ('sozlesme', 'kontrat', 'contract', 'anl', 'sözlesme')):
                contract_col = k
            elif scope_col is None and any(x in kn for x in
                    ('scope', 'kapsam', 'tanim', 'work', 'is')):
                scope_col = k
            elif project_col is None and any(x in kn for x in
                    ('proje', 'project', 'grup', 'group')):
                project_col = k
        # Pass 2 – value pattern: find column whose values look like 20xx-xx-xx
        if contract_col is None:
            for k in sample.keys():
                hits = sum(1 for r in rows[:10] if CONTRACT_RE.match(str(r.get(k) or '')))
                if hits >= max(1, len(rows[:10]) // 2):
                    contract_col = k
                    break
        return employer_col, contract_col, scope_col, project_col

    # ── Build from Hakedis Excel + DB hybrid ────────────────────────────────
    # If Hakedis Excel is uploaded:
    #   contracts come from its SÖZLEŞME NO column (these match DB İşveren-Sözleşme No)
    #   company grouping comes from its FİRMA column (= DB Company)
    # Otherwise: contracts/scopes always come from DB records.
    # Employers (İşveren/client) come from Access if uploaded, otherwise DB.
    # ─────────────────────────────────────────────────────────────────────────

    # Step 1: try to get contracts from Hakedis Excel
    hakedis_contracts_map = {}   # { company: [contractNo, ...] }
    hakedis_rows = _read_excel_rows('hakedis')
    if hakedis_rows:
        # Detect SÖZLEŞME NO and FİRMA columns
        sozlesme_col = None
        firma_col    = None
        if hakedis_rows:
            s = hakedis_rows[0]
            for k in s.keys():
                kn = _norm_col(k)
                if sozlesme_col is None and 'sozlesme' in kn:
                    sozlesme_col = k
                if firma_col is None and ('firma' in kn or 'company' in kn or 'sirket' in kn):
                    firma_col = k
        if sozlesme_col:
            for row in hakedis_rows:
                con = str(row.get(sozlesme_col, '') or '').strip()
                if not con or con in ('nan', 'None'):
                    continue
                firm = str(row.get(firma_col, '') or '').strip() if firma_col else ''
                if not firm or firm in ('nan', 'None'):
                    firm = ''
                if firm not in hakedis_contracts_map:
                    hakedis_contracts_map[firm] = set()
                hakedis_contracts_map[firm].add(con)
        # Convert sets to sorted lists
        hakedis_contracts_map = {k: sorted(v) for k, v in hakedis_contracts_map.items()}

    # Step 1b: get employers from Access (if uploaded)
    access_employers = []
    if _access_store and _access_store.get('contractsFromAccess'):
        rows = _access_store['contractsFromAccess']
        emp_col, _, _, _ = _detect_columns(rows)
        if emp_col:
            access_employers = sorted({
                str(r.get(emp_col) or '').strip()
                for r in rows
                if str(r.get(emp_col) or '').strip() not in ('', 'None')
            })

    # Step 2: build contracts/scopes/projects from DB records
    records = _load_records()
    employers_set  = set()
    db_contracts_map = defaultdict(set)  # company → set of contracts
    scopes_map     = defaultdict(set)
    projects_map   = defaultdict(set)

    for r in records:
        emp = _safe_str(r.get('Company', ''))
        con = _safe_str(r.get('İşveren- Sözleşme No', ''))
        sco = _safe_str(r.get('Scope', ''))
        pro = _safe_str(r.get('Projects/Group', '') or r.get('Projects', ''))
        if emp:
            employers_set.add(emp)
            if con:
                db_contracts_map[emp].add(con)
                if sco: scopes_map[con].add(sco)
                if pro: projects_map[con].add(pro)

    # Decide final contracts map:
    # If Hakedis Excel has FIRMA column grouping → use it per company
    # If Hakedis Excel has data but no FIRMA grouping → use it for ALL companies (empty key)
    # Otherwise fall back to DB
    if hakedis_contracts_map:
        if '' in hakedis_contracts_map and len(hakedis_contracts_map) == 1:
            # No FIRMA grouping found — broadcast to all DB companies
            all_contracts = hakedis_contracts_map['']
            contracts_map = {c: sorted(all_contracts) for c in employers_set}
        else:
            # FIRMA-keyed map — merge with DB company names
            # Try direct match first; then case-insensitive
            contracts_map = {}
            for company in employers_set:
                c_lower = company.lower()
                matched = (hakedis_contracts_map.get(company)
                           or next((v for k, v in hakedis_contracts_map.items()
                                    if k.lower() == c_lower), None)
                           or db_contracts_map.get(company, set()))
                contracts_map[company] = sorted(set(matched))
            # Also include any FIRMA entries that didn't match a DB company
            for k, v in hakedis_contracts_map.items():
                if k and k not in contracts_map:
                    contracts_map[k] = sorted(v)
    else:
        contracts_map = {e: sorted(c) for e, c in db_contracts_map.items()}

    # Prefer Access employer list if richer; otherwise fall back to DB companies
    employers_final = access_employers if access_employers else sorted(employers_set)
    source = 'hakedis+db' if hakedis_contracts_map else ('access+db' if access_employers else 'db')

    # ── Apply Year + Employer keyword filtering ───────────────────────────────
    contract_kw, _ = _employer_meta(cascade_employer)

    def _contract_matches(con):
        con_str = str(con)
        # Year filter
        if cascade_year:
            year_int = int(cascade_year) if cascade_year.isdigit() else None
            if year_int:
                # 2026 spillover: include 2025 AND 2026 contracts
                if year_int == 2026:
                    if not (con_str.startswith('2025') or con_str.startswith('2026')):
                        return False
                else:
                    if not con_str.startswith(cascade_year):
                        return False
        # Employer keyword filter
        if contract_kw:
            if contract_kw.upper() not in con_str.upper():
                return False
        return True

    if cascade_year or contract_kw:
        filtered_contracts_map = {}
        for comp, cons in contracts_map.items():
            filtered = [c for c in cons if _contract_matches(c)]
            filtered_contracts_map[comp] = filtered
        contracts_map = filtered_contracts_map

    return jsonify({
        'success':    True,
        'source':     source,
        'employers':  employers_final,
        'contracts':  contracts_map,
        'scopes':     {c: sorted(s) for c, s in scopes_map.items()},
        'projects':   {c: sorted(p) for c, p in projects_map.items()},
        'detectedCols': None,
        'rows': [],
    })


# ---------------------------------------------------------------------------
# Access helper: parse tables + VBA
# ---------------------------------------------------------------------------

def _parse_access_file(file_path: str) -> dict:
    """
    Parse an Access DB (.accdb / .mdb) and return:
      - tables: schema + sample rows
      - queries: saved query names + SQL text
      - macros: Access macro objects (non-VBA)
      - vbaModules: VBA code modules (Sub/Function/Class/Form)
      - forms: form names and their command-button click actions
      - contractsFromAccess: rows from any Sözleşme / Contract table
      - fieldMapping: DB field ↔ Access column cross-reference
    """
    tables_info   = []
    queries_info  = []
    macros_info   = []   # Access macro objects (not VBA)
    vba_modules   = []   # VBA code modules
    forms_info    = []   # Access form objects
    contracts_from_access = []
    errors = []

    # ── 1. ODBC: schema + data ────────────────────────────────────────────────
    conn = None
    try:
        import pyodbc
        conn_str = (
            r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
            f'DBQ={file_path};'
        )
        conn = pyodbc.connect(conn_str, autocommit=True)
        cursor = conn.cursor()

        # ── 1a. User tables ──────────────────────────────────────────────────
        raw_tables = [row.table_name for row in cursor.tables(tableType='TABLE')
                      if not row.table_name.startswith('MSys')]
        for tname in sorted(raw_tables):
            try:
                cols = []
                for col in cursor.columns(table=tname):
                    cols.append({
                        'name':     col.column_name,
                        'type':     col.type_name,
                        'nullable': col.nullable == 1,
                        'size':     col.column_size,
                    })
                try:
                    cursor.execute(f'SELECT TOP 5 * FROM [{tname}]')
                    col_names  = [d[0] for d in cursor.description]
                    sample_rows = []
                    for row in cursor.fetchall():
                        sample_rows.append({
                            col_names[i]: (str(v)[:120] if v is not None else None)
                            for i, v in enumerate(row)
                        })
                except Exception:
                    col_names   = [c['name'] for c in cols]
                    sample_rows = []

                tables_info.append({
                    'name':       tname,
                    'columns':    cols,
                    'colNames':   col_names if 'col_names' in dir() else [c['name'] for c in cols],
                    'sample':     sample_rows,
                    'rowCount':   None,
                })

                # ── Detect Sözleşme / Contract table ────────────────────────
                norm = re.sub(r'[^a-z0-9]', '', tname.lower())
                if any(k in norm for k in ('sozlesme', 'sozleme', 'contract', 'kontrat', 'anlaşma')):
                    try:
                        cursor.execute(f'SELECT * FROM [{tname}]')
                        c_names = [d[0] for d in cursor.description]
                        for row in cursor.fetchall():
                            contracts_from_access.append({
                                c_names[i]: (str(v)[:200] if v is not None else None)
                                for i, v in enumerate(row)
                            })
                    except Exception as ce:
                        errors.append(f'Contract table read {tname}: {ce}')

            except Exception as e:
                errors.append(f'Table {tname}: {e}')

        # ── 1b. Saved queries (Views) ────────────────────────────────────────
        raw_views = sorted(row.table_name for row in cursor.tables(tableType='VIEW'))

        # Try reading SQL text from MSysQueries
        sql_map = {}
        try:
            cursor.execute(
                "SELECT MSysObjects.Name, MSysQueries.Sql1 "
                "FROM MSysObjects INNER JOIN MSysQueries "
                "ON MSysObjects.ID = MSysQueries.ObjectId "
                "WHERE MSysObjects.Type = 5 ORDER BY MSysObjects.Name"
            )
            for row in cursor.fetchall():
                sql_map[row[0]] = row[1]
        except Exception:
            pass  # MSysQueries may be locked; fall back to name-only

        for vname in raw_views:
            queries_info.append({
                'name': vname,
                'sql':  sql_map.get(vname),
            })

        # ── 1c. Access Macros (non-VBA, Type = -32766) ──────────────────────
        try:
            cursor.execute(
                "SELECT Name FROM MSysObjects WHERE Type = -32766 ORDER BY Name"
            )
            for row in cursor.fetchall():
                macros_info.append({'name': row[0], 'type': 'macro'})
        except Exception:
            pass

        # ── 1d. Access Forms (Type = -32768) ────────────────────────────────
        try:
            cursor.execute(
                "SELECT Name FROM MSysObjects WHERE Type = -32768 ORDER BY Name"
            )
            for row in cursor.fetchall():
                forms_info.append({'name': row[0], 'buttons': []})
        except Exception:
            pass

    except Exception as e:
        errors.append(f'ODBC read failed: {e}')
    finally:
        try:
            if conn:
                conn.close()
        except Exception:
            pass

    # ── 2. VBA extraction ────────────────────────────────────────────────────
    # Access.Application COM automation requires an STA message pump.
    # Flask's thread pool doesn't provide one, causing "Visible property" DAO
    # errors.  Solution: spawn a dedicated subprocess which initialises COM in
    # its own STA context, reads all VBA modules, and returns JSON via stdout.
    #
    # Fallback: oletools for older .mdb files (OLE2 format).
    hakedis_keywords = ['hakedis', 'hakediş', 'payment', 'progres', 'pp', 'proje',
                        'period', 'sozlesme', 'sözleşme', 'hesap', 'rapor', 'excel']

    def _process_vba_module(module_name, vba_code):
        if not vba_code or not vba_code.strip():
            return
        subs = re.findall(
            r'(?m)^(?:Public |Private |Friend )?(?:Sub|Function)\s+(\w+)\s*\(',
            vba_code
        )
        click_handlers = re.findall(
            r'(?m)^(?:Public |Private )?Sub\s+(\w*(?:Click|_Click|Press|Tiklama)\w*)\s*\(',
            vba_code, re.IGNORECASE
        )
        relevant = any(kw in vba_code.lower() for kw in hakedis_keywords)

        form_prefix = module_name.split('.')[-1]
        for fi, frm in enumerate(forms_info):
            frm_norm = re.sub(r'[^a-z0-9]', '', frm['name'].lower())
            mod_norm = re.sub(r'[^a-z0-9]', '', form_prefix.lower())
            if frm_norm in mod_norm or mod_norm in frm_norm:
                forms_info[fi]['buttons'].extend(click_handlers)
                forms_info[fi]['module'] = module_name
                break

        vba_modules.append({
            'module':        module_name,
            'subs':          subs,
            'clickHandlers': click_handlers,
            'relevant':      relevant,
            'linesCount':    len(vba_code.splitlines()),
            'code':          vba_code if len(vba_code) < 16000 else vba_code[:16000] + '\n... [truncated]',
        })

    vba_extracted = False

    # ── 2a. Subprocess COM approach (works for .accdb and modern .mdb) ───────
    try:
        import subprocess, sys as _sys
        reader_script = os.path.join(
            os.path.dirname(__file__), '..', 'utils', 'access_vba_reader.py'
        )
        reader_script = os.path.normpath(reader_script)

        proc = subprocess.run(
            [_sys.executable, reader_script, file_path],
            capture_output=True,
            text=True,
            timeout=120,   # 2 min max; large databases can be slow
        )

        if proc.returncode == 0 and proc.stdout.strip():
            raw_modules = json.loads(proc.stdout)
            for m in raw_modules:
                _process_vba_module(m['name'], m['code'])
            vba_extracted = True
        else:
            stderr_msg = proc.stderr.strip() if proc.stderr else 'no output'
            errors.append(f'VBA subprocess error: {stderr_msg}')

    except Exception as e:
        errors.append(f'VBA subprocess launch failed: {e}')

    # ── 2b. oletools fallback (older .mdb OLE2 format only) ──────────────────
    if not vba_extracted:
        try:
            from oletools.olevba import VBA_Parser
            parser = VBA_Parser(file_path)
            has_vba = False
            try:
                has_vba = parser.detect_vba_macros()
            except Exception:
                has_vba = False

            if has_vba:
                for entry in parser.extract_macros():
                    try:
                        _, stream_path, vba_filename, vba_code = entry
                    except (ValueError, TypeError):
                        continue
                    module_name = (vba_filename or stream_path or 'Unknown').strip()
                    _process_vba_module(module_name, vba_code)
                vba_extracted = True
            parser.close()
        except Exception as e:
            errors.append(f'oletools VBA fallback also failed: {e}')

    # ── 3. Cross-reference Access columns with DatabaseRecord fields ─────────
    DB_FIELDS = [
        'Name Surname', 'Discipline', '(Week / Month)', 'Company',
        'Projects/Group', 'Scope', 'Projects', 'TOTAL MH',
        'North/ South', 'İşveren- Sözleşme No', 'İşveren- Hakediş (USD)',
        'General Total Cost (USD)', 'İşveren- Hakediş No',
        'AP-CB / Subcon', 'LS/Unit Rate',
    ]
    access_col_names = {
        col['name'].lower()
        for t in tables_info for col in t['columns']
    }
    field_mapping = []
    for db_f in DB_FIELDS:
        match = None
        db_normalized = re.sub(r'[^a-z0-9]', '', db_f.lower())
        for ac in access_col_names:
            ac_normalized = re.sub(r'[^a-z0-9]', '', ac)
            if db_normalized == ac_normalized or db_normalized in ac_normalized or ac_normalized in db_normalized:
                match = ac
                break
        field_mapping.append({'dbField': db_f, 'accessField': match, 'matched': match is not None})

    return {
        'tables':               tables_info,
        'queries':              queries_info,
        'macros':               macros_info,
        'vbaModules':           vba_modules,
        'forms':                forms_info,
        'contractsFromAccess':  contracts_from_access,
        'fieldMapping':         field_mapping,
        'errors':               errors,
        'summary': {
            'tableCount':       len(tables_info),
            'queryCount':       len(queries_info),
            'macroCount':       len(macros_info),
            'moduleCount':      len(vba_modules),
            'formCount':        len(forms_info),
            'contractCount':    len(contracts_from_access),
            'relevantModules':  [m['module'] for m in vba_modules if m['relevant']],
        },
    }


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def _build_excel(rows, totals, meta, person_detail=None):
    """Build and return an Excel workbook as bytes matching the official template."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Hakedis Raporu'

    # ── Colours ──────────────────────────────────────────────────────────────
    NAVY       = '1F3864'
    COL_HDR    = '4472C4'   # blue column header background
    GREY_ROW   = 'F2F2F2'
    WHITE      = 'FFFFFF'
    TOTAL_BG   = 'FFF2CC'   # light yellow totals
    SECTION_BG = 'D6DCE4'   # light grey section sub-headers

    cur_label = meta.get('currency', 'TL')
    # Currency format with symbol
    if cur_label == 'USD' or cur_label == '$':
        HAK_FMT = '#,##0.00 "$"'
    elif cur_label == 'EUR' or cur_label == '€':
        HAK_FMT = '#,##0.00 "€"'
    else:  # TL or default
        HAK_FMT = '#,##0.00 "₺"'

    NC = 5   # number of columns: A=Work Def, B=PP No, C=Cumulative, D=Previous, E=This Period

    # ── Project name prefix encoder ───────────────────────────────────────────
    _PROJECT_PREFIX_MAP = [
        ('USTLUGA', '03'),   # longer prefixes first to avoid partial matches
        ('MNLG',    '02'),
        ('GBS',     '01'),
    ]
    def _encode_project(name):
        """Replace known project prefixes with numeric codes, e.g. USTLUGA → 03."""
        if not name:
            return name
        upper = name.upper()
        for prefix, code in _PROJECT_PREFIX_MAP:
            if upper.startswith(prefix):
                return code + name[len(prefix):]
        return name

    # ── Column widths ────────────────────────────────────────────────────────
    # Col1: longest project name ~34 chars; cols 2-5: compact for numbers + wrapped headers
    for i, w in enumerate([34, 13, 17, 19, 17], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Helper factories ─────────────────────────────────────────────────────
    def _fill(hex_col):
        return PatternFill('solid', fgColor=hex_col)

    def _border(style='thin', color='999999'):
        s = Side(style=style, color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    THIN  = _border()
    THICK = _border('medium', '000000')

    def _cell(row, col, value='', bold=False, italic=False, size=11,
              fg='000000', bg=None, halign='left', fmt=None, wrap=True):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, italic=italic, color=fg, size=size)
        c.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=wrap)
        if bg:
            c.fill = _fill(bg)
        if fmt:
            c.number_format = fmt
        return c

    def _merge(row, c1, c2, row2=None):
        ws.merge_cells(
            start_row=row, start_column=c1,
            end_row=row2 or row, end_column=c2
        )

    def _row_h(row, h):
        ws.row_dimensions[row].height = h

    # ── Build row by row ─────────────────────────────────────────────────────
    company  = meta.get('company', '')
    employer = meta.get('employer', '')
    r = 1  # current row pointer

    # ── Title ─────────────────────────────────────────────────────────────────
    _merge(r, 1, NC)
    c = _cell(r, 1,
              f'{company} Hakedis Raporu / Progress Payment',
              bold=True, size=13, fg=WHITE, bg=NAVY, halign='center')
    c.border = THIN
    _row_h(r, 32)
    r += 1

    # ── Info block ────────────────────────────────────────────────────────────
    info_pairs = [
        ('İŞVEREN / EMPLOYER:',          employer),
        ('PROJECİ / PROJECT DESIGNER:',  company),
        ('SÖZLEŞME NO / CONTRACT NO:',   meta.get('contractNo', '')),
        ('DÖNEM / PERIOD:',              meta.get('periodText') or
                                         f"{meta.get('startDate','')} — {meta.get('endDate','')}"),
        ('TARİH / DATE OF ISSUE:',       meta.get('reportDate', '')),
    ]
    for label, val in info_pairs:
        lc = _cell(r, 1, label, bold=True)
        lc.border = THIN
        _merge(r, 2, NC)
        vc = _cell(r, 2, val, wrap=True)
        vc.border = THIN
        # Apply border to all merged cells in cols 3-NC
        for col in range(3, NC + 1):
            ws.cell(row=r, column=col).border = THIN
        # Dynamic height: merged value cols ~65 printable chars wide at size 11
        val_str = str(val) if val else ''
        lines = max(1, -(-len(val_str) // 65))
        _row_h(r, max(20, lines * 18))
        r += 1

    r += 1  # blank spacer

    # ── SECTION A ─────────────────────────────────────────────────────────────
    _merge(r, 1, NC)
    c = _cell(r, 1, 'A - Yapılan İşin Bedeli / Actual Cost for Work Performed',
              bold=True, size=12, fg=WHITE, bg=NAVY, halign='left')
    c.border = THIN
    _row_h(r, 24)
    r += 1

    # Table column headers
    col_headers = [
        'İş Tanımı / Work Definition ',
        'Hakediş No / Progress Payment No',
        f'Kümülatif Hakediş Tutarı / Cumulative Progress Payment ({cur_label})',
        f'Bir Önceki Kümülatif Hakediş Tutarı / Previous Period Progress Payment ({cur_label})',
        f'Bu Dönem Hakediş Tutarı / This Period Progress Payment ({cur_label})',
    ]
    _row_h(r, 80)
    for col, hdr in enumerate(col_headers, start=1):
        c = _cell(r, col, hdr, bold=True, fg=WHITE, bg=COL_HDR,
                  halign='center', wrap=True)
        c.border = THIN
    r += 1

    # Data rows
    for i, row_data in enumerate(rows):
        row_bg = WHITE if i % 2 == 0 else GREY_ROW
        _row_h(r, 20)
        c1 = _cell(r, 1, _encode_project(row_data['project']), bg=row_bg, halign='left');  c1.border = THIN
        c2 = _cell(r, 2,
                   f"PP-{row_data['ppNo']}" if row_data['ppNo'] else '',
                   bg=row_bg, halign='center');                            c2.border = THIN
        c3 = _cell(r, 3, row_data['cumulativeHakedis'],
                   bg=row_bg, halign='center', fmt=HAK_FMT);              c3.border = THIN
        c4 = _cell(r, 4, row_data['previousHakedis'],
                   bg=row_bg, halign='center', fmt=HAK_FMT);              c4.border = THIN
        c5 = _cell(r, 5, row_data['currentHakedis'],
                   bg=row_bg, halign='center', fmt=HAK_FMT);              c5.border = THIN
        r += 1

    # Sub-total row (dash label, bold values)
    _row_h(r, 20)
    for col in range(1, NC + 1):
        ws.cell(row=r, column=col).border = THIN
    _cell(r, 1, '-')
    _cell(r, 3, totals['cumulativeHakedis'], bold=True, halign='center', fmt=HAK_FMT).border = THIN
    _cell(r, 4, totals['previousHakedis'],  bold=True, halign='center', fmt=HAK_FMT).border = THIN
    _cell(r, 5, totals['currentHakedis'],   bold=True, halign='center', fmt=HAK_FMT).border = THIN
    r += 1

    # Calculate VAT (20%)
    vat_cumulative = totals['cumulativeHakedis'] * 0.20
    vat_previous = totals['previousHakedis'] * 0.20
    vat_current = totals['currentHakedis'] * 0.20

    # KDV / VAT 20%
    _row_h(r, 20)
    _cell(r, 1, 'KDV / VAT 20%', bold=True).border = THIN
    for col in range(2, NC + 1):
        ws.cell(row=r, column=col).border = THIN
    _cell(r, 3, vat_cumulative, halign='center', fmt=HAK_FMT).border = THIN
    _cell(r, 4, vat_previous, halign='center', fmt=HAK_FMT).border = THIN
    _cell(r, 5, vat_current, halign='center', fmt=HAK_FMT).border = THIN
    r += 1

    # Calculate totals with VAT
    total_with_vat_cumulative = totals['cumulativeHakedis'] + vat_cumulative
    total_with_vat_previous = totals['previousHakedis'] + vat_previous
    total_with_vat_current = totals['currentHakedis'] + vat_current

    # Toplam / Total
    _row_h(r, 24)
    _cell(r, 1, 'Toplam / Total', bold=True, bg=TOTAL_BG).border = THIN
    ws.cell(row=r, column=2).fill = _fill(TOTAL_BG); ws.cell(row=r, column=2).border = THIN
    _cell(r, 3, total_with_vat_cumulative, bold=True, bg=TOTAL_BG, halign='center', fmt=HAK_FMT).border = THIN
    _cell(r, 4, total_with_vat_previous, bold=True, bg=TOTAL_BG, halign='center', fmt=HAK_FMT).border = THIN
    _cell(r, 5, total_with_vat_current, bold=True, bg=TOTAL_BG, halign='center', fmt=HAK_FMT).border = THIN
    r += 1

    r += 1  # spacer

    # ── SECTION B: KESINTILER ─────────────────────────────────────────────────
    _merge(r, 1, NC)
    c = _cell(r, 1, 'B - Kesintiler / Deductions',
              bold=True, size=12, fg=WHITE, bg=NAVY, halign='left')
    c.border = THIN
    _row_h(r, 24)
    r += 1

    # Section B column headers
    _merge(r, 1, 2)
    ws.cell(row=r, column=1).fill = _fill(COL_HDR); ws.cell(row=r, column=1).border = THIN
    for col, lbl in [(3, 'Cumulative'), (4, 'Previous Period'), (5, 'This Period')]:
        _cell(r, col, lbl, bold=True, fg=WHITE, bg=COL_HDR, halign='center').border = THIN
    _row_h(r, 22)
    r += 1

    # Kesintiler row
    _merge(r, 1, 2)
    _cell(r, 1, 'Kesintiler / Deductions').border = THIN
    ws.cell(row=r, column=2).border = THIN
    for col in (3, 4, 5):
        _cell(r, col, 0, halign='center', fmt=HAK_FMT).border = THIN
    _row_h(r, 20)
    r += 1

    # Kesintiler Toplamı
    _merge(r, 1, 2)
    _cell(r, 1, 'Kesintiler Toplamı / Deductions Total', bold=True, bg=TOTAL_BG).border = THIN
    ws.cell(row=r, column=2).fill = _fill(TOTAL_BG); ws.cell(row=r, column=2).border = THIN
    for col in (3, 4, 5):
        _cell(r, col, 0, bold=True, bg=TOTAL_BG, halign='center', fmt=HAK_FMT).border = THIN
    _row_h(r, 20)
    r += 1

    r += 1  # spacer

    # ── SECTION C: NET PAYMENT ────────────────────────────────────────────────
    _merge(r, 1, NC)
    c = _cell(r, 1, 'C - Net Hakediş Tutarı / Net Progress Payment (C=A-B)',
              bold=True, size=12, fg=WHITE, bg=NAVY, halign='left')
    c.border = THIN
    _row_h(r, 24)
    r += 1

    # Net row (label spans cols 1-4, value in col 5)
    _merge(r, 1, 4)
    _cell(r, 1, 'Net Hakediş Tutarı / Net Progress Payment Amount').border = THIN
    for col in (2, 3, 4):
        ws.cell(row=r, column=col).border = THIN
    _cell(r, 5, total_with_vat_current, bold=True, halign='center', fmt=HAK_FMT).border = THIN
    _row_h(r, 20)
    r += 1

    # Toplam / Total Amount
    _merge(r, 1, 4)
    _cell(r, 1, 'Toplam / Total Amount', bold=True, bg=TOTAL_BG).border = THIN
    for col in (2, 3, 4):
        ws.cell(row=r, column=col).fill = _fill(TOTAL_BG)
        ws.cell(row=r, column=col).border = THIN
    _cell(r, 5, total_with_vat_cumulative, bold=True, bg=TOTAL_BG, halign='center', fmt=HAK_FMT).border = THIN
    _row_h(r, 20)
    r += 1

    r += 2  # spacer before footer

    # ── FOOTER ────────────────────────────────────────────────────────────────
    _merge(r, 1, 2)
    c = ws.cell(row=r, column=1, value='PROJECİ / PROJECT DESIGNER')
    c.font = Font(bold=True, color=NAVY, size=11)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    _merge(r, 4, 5)
    c = ws.cell(row=r, column=4, value='İŞVEREN / EMPLOYER')
    c.font = Font(bold=True, color=NAVY, size=11)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    _row_h(r, 18)
    r += 1

    _merge(r, 1, 2)
    c = ws.cell(row=r, column=1, value=company)
    c.font = Font(size=11)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    _merge(r, 4, 5)
    c = ws.cell(row=r, column=4, value=employer)
    c.font = Font(size=11)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    _row_h(r, 48)

    # Freeze under column headers (info block = 7 rows + 1 spacer = row 9)
    ws.freeze_panes = ws.cell(row=10, column=1)

    # Print settings: fit all columns on one page
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToPage  = True
    ws.page_setup.fitToWidth = 1   # 1 page wide
    ws.page_setup.fitToHeight = 0  # unlimited pages tall
    ws.page_setup.orientation = 'portrait'

    # Keep the old variables so Sheet 2 code still works
    MH_FMT = '#,##0.00'
    # Aliases for Sheet 2 helpers
    fill = _fill
    thin_border = lambda: THIN
    bold_border = lambda: THICK
    BLUE_DARK    = NAVY
    GREY         = GREY_ROW
    TOTAL_YELLOW = TOTAL_BG

    # ================================================================
    # Sheet 2 – Personel Detayı
    # ================================================================
    if person_detail:
        ws2 = wb.create_sheet('Personel Detayı')

        # Header row
        det_headers = [
            'Personel',
            'Proje',
            'Bu Dönem Toplam Çalışma',
            'Bu Dönem Hakediş Tutarı / This Period Progress Payment',
            'Geçmiş Dönem Toplam Çalışma',
            'Bir Önceki Kümülatif Hakediş Tutarı / Previous Period Progress Payment',
            'Kümülatif Hakediş Tutarı / Cumulative Progress Payment',
            'Birim Fiyat',
            'Para Birimi',
        ]
        ws2.row_dimensions[1].height = 30
        for col, hdr in enumerate(det_headers, start=1):
            cell = ws2.cell(row=1, column=col, value=hdr)
            cell.font = Font(bold=True, color=WHITE, size=11)
            cell.fill = fill(BLUE_DARK)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border()

        det_row = 2
        # Iterate projects in same sorted order as sheet 1
        sorted_projects = sorted(person_detail.keys())
        for proj in sorted_projects:
            persons = person_detail[proj]
            # Sort persons alphabetically
            for pname in sorted(persons.keys()):
                pd = persons[pname]
                row_fill = fill(WHITE) if det_row % 2 == 0 else fill(GREY)
                cur_mh    = pd['cur_mh']
                cur_tot   = pd['cur_total']
                prev_mh   = pd['prev_mh']
                prev_tot  = pd['prev_total']
                cumul     = cur_tot + prev_tot
                rate      = pd['rate']
                currency  = pd['currency'] or ''

                vals = [
                    pname,
                    _encode_project(proj),
                    cur_mh, cur_tot,
                    prev_mh, prev_tot,
                    cumul,
                    rate, currency,
                ]
                fmts = [
                    None,
                    None,
                    MH_FMT, HAK_FMT,
                    MH_FMT, HAK_FMT,
                    HAK_FMT,
                    '#,##0.00', None,
                ]
                for col, (val, fmt) in enumerate(zip(vals, fmts), start=1):
                    c = ws2.cell(row=det_row, column=col, value=val)
                    c.fill = row_fill
                    c.border = thin_border()
                    c.alignment = Alignment(
                        horizontal='left' if col <= 2 else 'right',
                        vertical='center'
                    )
                    if fmt:
                        c.number_format = fmt
                det_row += 1

        # Totals row on sheet 2
        total_row_fill = fill(TOTAL_YELLOW)
        ws2.row_dimensions[det_row].height = 20
        ws2.cell(row=det_row, column=1, value='TOPLAM').font = Font(bold=True)
        ws2.cell(row=det_row, column=1).fill = total_row_fill
        ws2.cell(row=det_row, column=1).border = bold_border()
        ws2.cell(row=det_row, column=2).fill = total_row_fill
        ws2.cell(row=det_row, column=2).border = bold_border()
        for col, idx in [(3, 'cur_mh'), (4, 'cur_total'), (5, 'prev_mh'),
                         (6, 'prev_total'), (7, None)]:
            if idx:
                val = sum(pd[idx] for ps in person_detail.values() for pd in ps.values())
            else:  # cumulative
                val = sum(pd['cur_total'] + pd['prev_total']
                          for ps in person_detail.values() for pd in ps.values())
            fmt = MH_FMT if col in (3, 5) else HAK_FMT
            c = ws2.cell(row=det_row, column=col, value=val)
            c.font = Font(bold=True)
            c.fill = total_row_fill
            c.border = bold_border()
            c.number_format = fmt
            c.alignment = Alignment(horizontal='right', vertical='center')

        # Column widths sheet 2
        det_widths = [28, 28, 18, 32, 18, 38, 32, 14, 12]
        for i, w in enumerate(det_widths, start=1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        ws2.freeze_panes = ws2.cell(row=2, column=1)

        # Print settings: fit all columns on one page
        ws2.sheet_properties.pageSetUpPr.fitToPage = True
        ws2.page_setup.fitToPage  = True
        ws2.page_setup.fitToWidth = 1
        ws2.page_setup.fitToHeight = 0
        ws2.page_setup.orientation = 'portrait'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
